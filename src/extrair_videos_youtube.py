# -*- coding: utf-8 -*-
"""
Extract the complete list of videos from a YouTube channel and write an Excel
workbook with title, description, URL, publication date, and duration.
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

IDENTIFICADOR_DO_CANAL = "UCj203koB0VhNWF43yy1w8ug"
ENDERECO_DO_CANAL = "https://www.youtube.com/@ozlsp12"
ARQUIVO_DE_SAIDA = "data/videos_canal_youtube.xlsx"


def coletar_por_interface_oficial(chave, identificador_do_canal):
    from googleapiclient.discovery import build

    servico = build("youtube", "v3", developerKey=chave)
    canal = servico.channels().list(
        part="contentDetails",
        id=identificador_do_canal,
    ).execute()

    itens = canal.get("items", [])
    if not itens:
        raise RuntimeError("Canal não encontrado pela YouTube Data API.")

    lista_de_envios = itens[0]["contentDetails"]["relatedPlaylists"]["uploads"]

    registros = []
    token_da_pagina = None
    while True:
        resposta = servico.playlistItems().list(
            part="snippet,contentDetails",
            playlistId=lista_de_envios,
            maxResults=50,
            pageToken=token_da_pagina,
        ).execute()

        for item in resposta.get("items", []):
            resumo = item["snippet"]
            identificador = item["contentDetails"]["videoId"]
            registros.append({
                "identificador": identificador,
                "titulo": resumo.get("title", ""),
                "descricao": resumo.get("description", ""),
                "publicacao": item["contentDetails"].get("videoPublishedAt", ""),
                "link": "https://www.youtube.com/watch?v=" + identificador,
                "duracao": "",
            })

        token_da_pagina = resposta.get("nextPageToken")
        if not token_da_pagina:
            break

    indice = {registro["identificador"]: registro for registro in registros}
    identificadores = list(indice.keys())
    for inicio in range(0, len(identificadores), 50):
        bloco = identificadores[inicio:inicio + 50]
        resposta = servico.videos().list(
            part="contentDetails",
            id=",".join(bloco),
        ).execute()
        for item in resposta.get("items", []):
            indice[item["id"]]["duracao"] = converter_duracao_iso(
                item["contentDetails"]["duration"]
            )

    return registros


def converter_duracao_iso(texto):
    padrao = re.match(
        r"P(?:(\d+)D)?T(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", texto or ""
    )
    if not padrao:
        return ""
    dias, horas, minutos, segundos = (int(g or 0) for g in padrao.groups())
    horas += dias * 24
    return "{:02d}:{:02d}:{:02d}".format(horas, minutos, segundos)


def coletar_por_ytdlp(endereco_do_canal):
    from yt_dlp import YoutubeDL

    opcoes_rasas = {"quiet": True, "extract_flat": True, "skip_download": True}
    with YoutubeDL(opcoes_rasas) as extrator:
        arvore = extrator.extract_info(endereco_do_canal + "/videos", download=False)

    entradas = arvore.get("entries") or []
    identificadores = [entrada["id"] for entrada in entradas if entrada.get("id")]

    opcoes_completas = {"quiet": True, "skip_download": True}
    registros = []
    with YoutubeDL(opcoes_completas) as extrator:
        for posicao, identificador in enumerate(identificadores, start=1):
            print("  {}/{}".format(posicao, len(identificadores)), file=sys.stderr)
            dados = extrator.extract_info(
                "https://www.youtube.com/watch?v=" + identificador, download=False
            )
            carimbo = dados.get("upload_date") or ""
            registros.append({
                "identificador": identificador,
                "titulo": dados.get("title", ""),
                "descricao": dados.get("description", "") or "",
                "publicacao": (
                    "{}-{}-{}".format(carimbo[:4], carimbo[4:6], carimbo[6:8])
                    if len(carimbo) == 8 else ""
                ),
                "link": "https://www.youtube.com/watch?v=" + identificador,
                "duracao": formatar_segundos(dados.get("duration")),
            })
    return registros


def formatar_segundos(total):
    if not total:
        return ""
    total = int(total)
    return "{:02d}:{:02d}:{:02d}".format(
        total // 3600, (total % 3600) // 60, total % 60
    )


def gravar_planilha(registros, caminho):
    caminho = Path(caminho)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    cabecalhos = ["Número", "Título", "Descrição", "Link", "Publicação", "Duração"]
    larguras = [8, 60, 90, 45, 14, 12]

    livro = Workbook()
    planilha = livro.active
    planilha.title = "Videos"

    for coluna, (rotulo, largura) in enumerate(zip(cabecalhos, larguras), start=1):
        celula = planilha.cell(row=1, column=coluna, value=rotulo)
        celula.font = Font(name="Arial", bold=True)
        celula.alignment = Alignment(vertical="center")
        planilha.column_dimensions[get_column_letter(coluna)].width = largura

    for linha, registro in enumerate(registros, start=2):
        valores = [
            linha - 1,
            registro["titulo"],
            registro["descricao"],
            registro["link"],
            registro["publicacao"][:10],
            registro["duracao"],
        ]
        for coluna, valor in enumerate(valores, start=1):
            celula = planilha.cell(row=linha, column=coluna, value=valor)
            celula.font = Font(name="Arial")
            celula.alignment = Alignment(vertical="top", wrap_text=(coluna == 3))

    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = "A1:F{}".format(max(len(registros) + 1, 2))

    nota = livro.create_sheet("Notas")
    nota["A1"] = "Canal"
    nota["B1"] = ENDERECO_DO_CANAL
    nota["A2"] = "Identificador do canal"
    nota["B2"] = IDENTIFICADOR_DO_CANAL
    nota["A3"] = "Extração"
    nota["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    nota["A4"] = "Total de vídeos"
    nota["B4"] = len(registros)
    for linha in range(1, 5):
        nota.cell(row=linha, column=1).font = Font(name="Arial", bold=True)
        nota.cell(row=linha, column=2).font = Font(name="Arial")
    nota.column_dimensions["A"].width = 26
    nota.column_dimensions["B"].width = 50

    livro.save(caminho)


def principal():
    analisador = argparse.ArgumentParser()
    analisador.add_argument("--modo", choices=["api", "ytdlp"], default="ytdlp")
    analisador.add_argument(
        "--chave",
        default=os.environ.get("YOUTUBE_API_KEY"),
        help=(
            "chave da YouTube Data API (modo api); se omitida, usa a variável "
            "de ambiente YOUTUBE_API_KEY"
        ),
    )
    analisador.add_argument("--saida", default=ARQUIVO_DE_SAIDA)
    argumentos = analisador.parse_args()

    if argumentos.modo == "api":
        if not argumentos.chave:
            analisador.error(
                "o modo api exige --chave ou a variável de ambiente YOUTUBE_API_KEY"
            )
        registros = coletar_por_interface_oficial(
            argumentos.chave, IDENTIFICADOR_DO_CANAL
        )
    else:
        registros = coletar_por_ytdlp(ENDERECO_DO_CANAL)

    registros.sort(key=lambda registro: registro["publicacao"], reverse=True)
    gravar_planilha(registros, argumentos.saida)
    print("{} vídeos gravados em {}".format(len(registros), argumentos.saida))


if __name__ == "__main__":
    principal()
