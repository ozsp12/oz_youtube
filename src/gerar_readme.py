from pathlib import Path
from openpyxl import load_workbook

ARQUIVO_EXCEL = Path("data/videos_canal_youtube.xlsx")
ARQUIVO_README = Path("README.md")
NOME_PLANILHA = "Videos"
MARCADOR_INICIO = "<!-- YOUTUBE-VIDEOS:START -->"
MARCADOR_FIM = "<!-- YOUTUBE-VIDEOS:END -->"
TAMANHO_MAXIMO_DESCRICAO = 220


def limpar_markdown(valor):
    if valor is None:
        return ""
    texto = str(valor).strip().replace("|", r"\|")
    texto = texto.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    return " ".join(texto.split())


def resumir_descricao(texto, limite=TAMANHO_MAXIMO_DESCRICAO):
    texto = limpar_markdown(texto)
    if len(texto) <= limite:
        return texto
    return texto[: limite - 3].rstrip() + "..."


def carregar_videos(caminho_excel):
    livro = load_workbook(caminho_excel, data_only=True, read_only=True)
    planilha = livro[NOME_PLANILHA]
    cabecalhos = [limpar_markdown(celula.value) for celula in planilha[1]]

    videos = []
    for linha in planilha.iter_rows(min_row=2, values_only=True):
        registro = dict(zip(cabecalhos, linha))
        if registro.get("Título"):
            videos.append(registro)
    return videos


def gerar_tabela_markdown(videos):
    linhas = [
        "## Videos",
        "",
        f"Total cataloged videos: **{len(videos)}**.",
        "",
        "| # | Video | Description | Published | Duration |",
        "|---:|---|---|:---:|:---:|",
    ]

    for video in videos:
        numero = limpar_markdown(video.get("Número"))
        titulo = limpar_markdown(video.get("Título"))
        descricao = resumir_descricao(video.get("Descrição"))
        link = limpar_markdown(video.get("Link"))
        publicacao = limpar_markdown(video.get("Publicação"))
        duracao = limpar_markdown(video.get("Duração"))
        titulo_markdown = f"[{titulo}]({link})" if link else titulo
        linhas.append(
            f"| {numero} | {titulo_markdown} | {descricao} | {publicacao} | {duracao} |"
        )

    return "\n".join(linhas)


def atualizar_readme(conteudo_gerado):
    if ARQUIVO_README.exists():
        texto = ARQUIVO_README.read_text(encoding="utf-8")
    else:
        texto = "# oz_youtube\n"

    bloco = (
        f"{MARCADOR_INICIO}\n\n"
        f"{conteudo_gerado}\n\n"
        f"{MARCADOR_FIM}"
    )

    if MARCADOR_INICIO in texto and MARCADOR_FIM in texto:
        antes = texto.split(MARCADOR_INICIO, 1)[0]
        depois = texto.split(MARCADOR_FIM, 1)[1]
        texto = antes.rstrip() + "\n\n" + bloco + "\n\n" + depois.lstrip()
    else:
        texto = texto.rstrip() + "\n\n" + bloco + "\n"

    ARQUIVO_README.write_text(texto, encoding="utf-8")


def main():
    videos = carregar_videos(ARQUIVO_EXCEL)
    atualizar_readme(gerar_tabela_markdown(videos))
    print(f"README updated with {len(videos)} videos.")


if __name__ == "__main__":
    main()
